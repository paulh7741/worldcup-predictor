#!/usr/bin/env python3
"""
World Cup 2026 Predictor - Auto Scoring System
Fetches live results, calculates points, updates spreadsheet, emails standings.

Setup:
  1. Get a FREE API key at https://www.football-data.org/client/register
  2. For email: enable Gmail 2-Step Verification, then create an App Password at
     https://myaccount.google.com/apppasswords  (select "Mail" + your device)
  3. Edit wc_config.json with your API key and email app-password.
  4. Run:  python3 wc_predictor.py
  5. Schedule (cron every 2 hours):
       crontab -e
       0 */2 * * * cd /home/paul41admin/projects/testing/Data && python3 wc_predictor.py >> wc_predictor.log 2>&1
"""

import hashlib
import json
import os
import smtplib
import sys
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import openpyxl
import requests
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent
SPREADSHEET_PATH = SCRIPT_DIR / "World Cup Predictor MastervAIxlsx.xlsx"
STATE_FILE = SCRIPT_DIR / "wc_state.json"
CONFIG_FILE = SCRIPT_DIR / "wc_config.json"

DEFAULT_CONFIG = {
    "football_api_key": "",          # football-data.org free key
    "competition_code": "WC",        # Competition code (WC = FIFA World Cup)
    "email_sender": "",              # Your Gmail address
    "email_password": "",            # Gmail App Password (not your login password)
    "email_recipient": "paulh7741@gmail.com",
    "smtp_host": "smtp.gmail.com",
    "smtp_port": 587,
    "use_api": True,                 # Set false to score from Results sheet only
    "send_email_on_change": True,
}

# ── Team Name Normalisation ────────────────────────────────────────────────────
# Maps any known variant -> canonical spreadsheet form (UPPERCASE)
_ALIASES: dict[str, str] = {}

_RAW_ALIASES = {
    "CZECH":        ["czech republic", "czechia", "czech"],
    "CURACO":       ["curaçao", "curacao", "curaçao"],
    "AUSTRAILIA":   ["australia"],
    "IVORY COAST":  ["côte d'ivoire", "cote d'ivoire", "ivory coast", "cote divoire"],
    "NEW ZELAND":   ["new zealand"],
    "CONGO DR":     ["dr congo", "democratic republic of congo", "congo dr",
                     "congo, dr", "congo (dr)", "congo-kinshasa"],
    "BOSNIA-HER":   ["bosnia and herzegovina", "bosnia & herzegovina", "bosnia"],
    "CABO VERDE":   ["cape verde", "cabo verde"],
    "SOUTH AFRICA": ["south africa"],
    "SAUDI ARABIA": ["saudi arabia"],
    "KOREA":        ["korea republic", "south korea", "korea"],
    "COLUMBIA":     ["colombia", "columbia"],   # spreadsheet spells it COLUMBIA
    "USA":          ["usa", "united states", "united states of america"],
    "TURKEY":       ["turkey", "türkiye"],
    "NETHERLANDS":  ["netherlands", "holland"],
}

for canonical, variants in _RAW_ALIASES.items():
    for v in variants:
        _ALIASES[v.lower()] = canonical
    _ALIASES[canonical.lower()] = canonical


def normalise(name: str | None) -> str | None:
    if name is None:
        return None
    key = str(name).strip().lower()
    return _ALIASES.get(key, str(name).strip().upper())


def match_result(s1: int, s2: int) -> str:
    return "H" if s1 > s2 else ("A" if s1 < s2 else "D")


def score_prediction(ps1, ps2, as1, as2) -> int:
    """3 for correct score, 1 for correct result, 0 otherwise."""
    if any(v is None for v in (ps1, ps2, as1, as2)):
        return 0
    if int(ps1) == int(as1) and int(ps2) == int(as2):
        return 3
    if match_result(int(ps1), int(ps2)) == match_result(int(as1), int(as2)):
        return 1
    return 0


def make_key(t1: str | None, t2: str | None) -> str:
    return f"{normalise(t1)}|{normalise(t2)}"


def lookup_result(actual: dict, key: str):
    """Return (s1, s2) for team-order in key, or None.  Tries reversed key."""
    if key in actual:
        return actual[key]
    parts = key.split("|")
    if len(parts) == 2:
        rev = f"{parts[1]}|{parts[0]}"
        if rev in actual:
            a, b = actual[rev]
            return (b, a)
    return None


# ── Labels to skip ─────────────────────────────────────────────────────────────
_SKIP_L = {
    "TEAMS", "GROUP WINNER", "Name", "RULES:", "SCORING",
    "CORRECT SCORE", "CORRECT RESULT", "GROUP WINNERS", "OVERALL WINNER",
    "TOP GOAL SCORER",
    "1.   £15.00 ENTRY PAID PRIOR TO TOURNAMENT START",
    "2.   FORMS TO BE COMPLETED PRIOR TO TOURNAMENT START. ANY FORMS",
    "       RECEIVED  AFTER THE START WILL NOT BE ENTERED",
    "3.   WINNER WILL BE DECEIDED BY THE HIGHEST NUMBER OF POINTS TOTALED",
    "       IN THE EVENT OF A TIE THE WINNER WILL BE DECIDED BY THE TIE BREAKER",
    "4.   PLEASE COMPLETE THE FORM BY ENTERING THE SCORE FOR EACH MATCH,",
    "       GROUP WINNER, OVERALL WINNER, TOP GOAL SCORER AND THE TIE BREAK",
}
_SKIP_R = {
    "GROUP WINNER", "OVERALL WINNER", "TOP GOAL SCORER",
    "TIE BREAKER ", "TIE BREAKER", "TOTAL TOURNAMENT GOALS",
    "RULES:", "SCORING", "CORRECT SCORE", "CORRECT RESULT", "GROUP WINNERS",
}
_SCORING_VALS = {"3 POINTS", "1 POINT", "5 POINTS"}


def _is_short_str(v) -> bool:
    return isinstance(v, str) and len(v.strip()) <= 35


# ── Parse player sheet ─────────────────────────────────────────────────────────
def parse_player_sheet(ws, sheet_name: str) -> dict:
    """Extract all predictions from one player sheet."""
    O = 1  # Column offset (player sheets have leading None in col A)
    name = None
    match_preds: dict[str, tuple] = {}   # key -> (s1, s2)
    match_group_map: dict[str, str] = {} # key -> group letter
    group_winners: dict[str, str] = {}   # group letter -> team
    overall_winner = None
    top_scorer = None
    tie_breaker = None
    cur_lg = None   # current left group
    cur_rg = None   # current right group

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if len(row) < 12:
            continue
        r = list(row)

        # Name
        if r[O] == "Name":
            name = r[O + 2]
            continue
        if r[O] == "TEAMS":
            continue

        # Track group labels
        g = r[O + 4]
        if isinstance(g, str) and len(g) == 1 and g.isalpha():
            cur_lg = g
        g = r[O + 10]
        if isinstance(g, str) and len(g) == 1 and g.isalpha():
            cur_rg = g

        # Is this a right-side scoring-rule row? (e.g. 'OVERALL WINNER', '3 POINTS')
        right_is_rule = (r[O + 7] in _SCORING_VALS or
                         (r[O + 6] in _SKIP_R and r[O + 7] in _SCORING_VALS))

        # ── Special right-side predictions (check BEFORE group-winner continue) ──
        # These can share a row with 'GROUP WINNER' on the left (e.g. row 39 in player sheets)
        if r[O + 6] == "OVERALL WINNER" and r[O + 7] is None:
            overall_winner = normalise(r[O + 8])
        if r[O + 6] == "TOP GOAL SCORER" and r[O + 7] is None:
            top_scorer = str(r[O + 8]).strip() if r[O + 8] else None
        if r[O + 6] in ("TIE BREAKER ", "TIE BREAKER") and r[O + 7] is None:
            tie_breaker = r[O + 8]
        if r[O + 6] == "TOTAL TOURNAMENT GOALS" and r[O + 7] is None:
            tie_breaker = r[O + 8]

        # ── Group winner rows ──
        if r[O] == "GROUP WINNER":
            lw = r[O + 2]
            if lw is not None and cur_lg:
                group_winners[cur_lg] = normalise(str(lw))
            if not right_is_rule and r[O + 6] == "GROUP WINNER":
                rw = r[O + 8]
                if rw is not None and cur_rg:
                    group_winners[cur_rg] = normalise(str(rw))
            continue

        # ── Left-side match ──
        t1, t2 = r[O], r[O + 1]
        if (t1 not in _SKIP_L and isinstance(t1, str) and _is_short_str(t2) and
                r[O + 2] is not None and r[O + 3] is not None):
            try:
                k = make_key(t1, t2)
                match_preds[k] = (int(r[O + 2]), int(r[O + 3]))
                if cur_lg:
                    match_group_map[k] = cur_lg
            except (TypeError, ValueError):
                pass

        # ── Right-side match (only if not a rule row and team2 is short) ──
        rt1, rt2 = r[O + 6], r[O + 7]
        if (not right_is_rule and rt1 not in _SKIP_R and _is_short_str(rt1) and
                rt1 is not None and _is_short_str(rt2) and rt2 is not None and
                rt2 not in _SCORING_VALS and
                isinstance(r[O + 8], (int, float)) and
                isinstance(r[O + 9], (int, float))):
            try:
                k = make_key(rt1, rt2)
                match_preds[k] = (int(r[O + 8]), int(r[O + 9]))
                if cur_rg:
                    match_group_map[k] = cur_rg
            except (TypeError, ValueError):
                pass

    return {
        "sheet_name": sheet_name,
        "name": (str(name).strip() if name else None) or sheet_name.split(" - ", 1)[-1].strip(),
        "match_preds": match_preds,
        "match_group_map": match_group_map,
        "group_winners": group_winners,
        "overall_winner": overall_winner,
        "top_scorer": top_scorer,
        "tie_breaker": tie_breaker,
    }


# ── Parse Results sheet ────────────────────────────────────────────────────────
def parse_results_sheet(ws) -> tuple[dict, dict, str | None, str | None, int | None]:
    """Return (results, group_winners, overall_winner, top_scorer, total_goals)."""
    results: dict[str, tuple] = {}
    group_winners: dict[str, str] = {}
    overall_winner = None
    top_scorer = None
    total_goals = None
    cur_lg = None
    cur_rg = None

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if len(row) < 11:
            continue
        r = list(row)

        # Track groups
        if isinstance(r[4], str) and len(r[4]) == 1 and r[4].isalpha():
            cur_lg = r[4]
        if isinstance(r[10], str) and len(r[10]) == 1 and r[10].isalpha():
            cur_rg = r[10]

        # Group winner row
        if r[0] == "GROUP WINNER":
            if r[2] and cur_lg:
                group_winners[cur_lg] = normalise(str(r[2]))
            if r[6] == "GROUP WINNER" and r[8] and cur_rg:
                group_winners[cur_rg] = normalise(str(r[8]))
            continue

        # Special fields
        if r[6] == "OVERALL WINNER" and r[8]:
            overall_winner = normalise(str(r[8]))
        if r[6] == "TOP GOAL SCORER" and r[8]:
            top_scorer = str(r[8]).strip()
        if r[6] == "TOTAL TOURNAMENT GOALS" and r[8]:
            total_goals = r[8]

        # Left match
        t1, t2 = r[0], r[1]
        if (_is_short_str(t1) and t1 not in _SKIP_L and
                _is_short_str(t2) and t2 is not None and
                isinstance(r[2], (int, float)) and isinstance(r[3], (int, float))):
            results[make_key(t1, t2)] = (int(r[2]), int(r[3]))

        # Right match
        rt1, rt2 = r[6], r[7]
        if (_is_short_str(rt1) and rt1 not in _SKIP_R and
                _is_short_str(rt2) and rt2 is not None and
                isinstance(r[8], (int, float)) and isinstance(r[9], (int, float))):
            results[make_key(rt1, rt2)] = (int(r[8]), int(r[9]))

    return results, group_winners, overall_winner, top_scorer, total_goals


# ── Football-data.org API ──────────────────────────────────────────────────────
def fetch_api_results(api_key: str, competition_code: str = "WC"):
    """Fetch finished match results and group standings from football-data.org.
    Returns (results_dict, group_winners_dict) or (None, None) on error."""
    if not api_key:
        return None, None

    headers = {"X-Auth-Token": api_key}
    results: dict[str, tuple] = {}
    group_winners: dict[str, str] = {}

    try:
        # ── Finished matches ──
        url = f"https://api.football-data.org/v4/competitions/{competition_code}/matches"
        resp = requests.get(url, headers=headers, params={"status": "FINISHED"}, timeout=15)
        resp.raise_for_status()
        for m in resp.json().get("matches", []):
            home = normalise(m["homeTeam"].get("shortName") or m["homeTeam"].get("name"))
            away = normalise(m["awayTeam"].get("shortName") or m["awayTeam"].get("name"))
            ft = m.get("score", {}).get("fullTime", {})
            if ft.get("home") is not None and ft.get("away") is not None:
                results[f"{home}|{away}"] = (int(ft["home"]), int(ft["away"]))

        # ── Standings (group winners = position-1 team per group) ──
        url2 = f"https://api.football-data.org/v4/competitions/{competition_code}/standings"
        resp2 = requests.get(url2, headers=headers, timeout=15)
        resp2.raise_for_status()
        for standing in resp2.json().get("standings", []):
            if standing.get("type") == "TOTAL":
                raw_group = standing.get("group", "")   # e.g. "GROUP_A"
                group_letter = raw_group.replace("GROUP_", "").strip()
                table = standing.get("table", [])
                if table:
                    played_rows = [t for t in table if t.get("playedGames", 0) > 0]
                    if played_rows:
                        winner_name = (played_rows[0]["team"].get("shortName") or
                                       played_rows[0]["team"].get("name"))
                        if group_letter:
                            group_winners[group_letter] = normalise(winner_name)

        print(f"  API: {len(results)} finished matches, {len(group_winners)} group leaders fetched.")
        return results, group_winners

    except requests.HTTPError as e:
        print(f"  API HTTP error: {e.response.status_code} {e.response.text[:200]}", file=sys.stderr)
        return None, None
    except Exception as e:
        print(f"  API error: {e}", file=sys.stderr)
        return None, None


# ── Update Results sheet with live data ───────────────────────────────────────
def update_results_sheet(ws, api_results: dict, api_gw: dict):
    """Write API results back into the Results sheet."""
    # Build cell map: match_key -> (row_num, score1_col, score2_col)  [1-based cols]
    cell_map: dict[str, tuple] = {}
    cur_lg = cur_rg = None

    for row_num, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True), start=3):
        if len(row) < 11:
            continue
        r = list(row)

        if isinstance(r[4], str) and len(r[4]) == 1 and r[4].isalpha():
            cur_lg = r[4]
        if isinstance(r[10], str) and len(r[10]) == 1 and r[10].isalpha():
            cur_rg = r[10]

        if r[0] == "GROUP WINNER":
            # Store group-winner row locations
            if cur_lg:
                cell_map[f"GW_LEFT|{cur_lg}"] = (row_num, 3, None)     # Col C
            if r[6] == "GROUP WINNER" and cur_rg:
                cell_map[f"GW_RIGHT|{cur_rg}"] = (row_num, 9, None)    # Col I
            continue

        # Left match
        t1, t2 = r[0], r[1]
        if _is_short_str(t1) and t1 not in _SKIP_L and _is_short_str(t2) and t2:
            cell_map[make_key(t1, t2)] = (row_num, 3, 4)   # Cols C, D

        # Right match
        rt1, rt2 = r[6], r[7]
        if _is_short_str(rt1) and rt1 not in _SKIP_R and _is_short_str(rt2) and rt2:
            cell_map[make_key(rt1, rt2)] = (row_num, 9, 10)  # Cols I, J

    # Write match scores
    for key, (s1, s2) in api_results.items():
        loc = None
        if key in cell_map:
            loc = cell_map[key]
            ws.cell(row=loc[0], column=loc[1], value=s1)
            ws.cell(row=loc[0], column=loc[2], value=s2)
        else:
            parts = key.split("|")
            if len(parts) == 2:
                rev = f"{parts[1]}|{parts[0]}"
                if rev in cell_map:
                    loc = cell_map[rev]
                    ws.cell(row=loc[0], column=loc[1], value=s2)   # swap
                    ws.cell(row=loc[0], column=loc[2], value=s1)

    # Write group winners
    for group, winner in api_gw.items():
        lk = f"GW_LEFT|{group}"
        rk = f"GW_RIGHT|{group}"
        if lk in cell_map:
            ws.cell(row=cell_map[lk][0], column=cell_map[lk][1], value=winner)
        if rk in cell_map:
            ws.cell(row=cell_map[rk][0], column=cell_map[rk][1], value=winner)


# ── Score calculation ──────────────────────────────────────────────────────────
def completed_groups(players: list, actual: dict) -> set:
    """Return set of group letters where all 6 matches are finished."""
    # Build group -> set of match keys from the first player sheet
    group_matches: dict[str, set] = {}
    if not players:
        return set()
    for key in players[0]["match_preds"]:
        # We need to know which group each match belongs to.
        # Re-derive from the first player's group_winners keys and match predictions.
        pass

    # Better: use the group tracker built during parsing.
    # Each player sheet stores group->winner; we need group->match keys.
    # Rebuild by scanning player 0's sheet predictions alongside group info.
    # Since we stored match_preds as flat dict, group membership is implicit.
    # Simplest reliable approach: count how many of a group's 6 expected matches
    # appear in actual results, using any player's predictions as the fixture list.
    p0 = players[0]

    # Build group -> list of match keys using group_match_map stored during parsing
    # (We'll add this in parse_player_sheet below — for now derive from match count)
    # Each group has exactly 6 matches (4 teams, round-robin)
    group_match_keys: dict[str, list] = {}
    for key in p0.get("match_group_map", {}):
        grp = p0["match_group_map"][key]
        group_match_keys.setdefault(grp, []).append(key)

    done = set()
    for grp, keys in group_match_keys.items():
        if len(keys) == 6 and all(lookup_result(actual, k) is not None for k in keys):
            done.add(grp)
    return done


def calculate_scores(players: list, actual: dict, actual_gw: dict,
                     actual_ow, actual_ts, actual_goals) -> list:
    """Return sorted list of player score dicts."""
    finished_groups = completed_groups(players, actual)
    scores = []

    for p in players:
        m_pts = 0
        for key, (ps1, ps2) in p["match_preds"].items():
            res = lookup_result(actual, key)
            if res:
                m_pts += score_prediction(ps1, ps2, res[0], res[1])

        gw_pts = 0
        for grp, pred_winner in p["group_winners"].items():
            # Only award group winner points once all 6 games in that group are done
            if grp not in finished_groups:
                continue
            if actual_gw.get(grp) and pred_winner:
                if normalise(pred_winner) == normalise(actual_gw[grp]):
                    gw_pts += 3

        ow_pts = 0
        if actual_ow and p["overall_winner"]:
            if normalise(p["overall_winner"]) == normalise(actual_ow):
                ow_pts = 3

        ts_pts = 0
        if actual_ts and p["top_scorer"]:
            a = actual_ts.strip().lower()
            pred = p["top_scorer"].strip().lower()
            if pred == a or pred in a or a in pred:
                ts_pts = 5

        total = m_pts + gw_pts + ow_pts + ts_pts

        tb = p["tie_breaker"]
        tb_diff = None
        if tb is not None and actual_goals is not None:
            try:
                tb_diff = abs(int(tb) - int(actual_goals))
            except (TypeError, ValueError):
                pass

        scores.append({
            "name": p["name"],
            "sheet_name": p["sheet_name"],
            "total": total,
            "match_pts": m_pts,
            "gw_pts": gw_pts,
            "ow_pts": ow_pts,
            "ts_pts": ts_pts,
            "overall_winner": p["overall_winner"],
            "top_scorer": p["top_scorer"],
            "tie_breaker": tb,
            "_tb_diff": tb_diff,
        })

    # Sort: highest total, then smallest tiebreaker distance
    scores.sort(key=lambda s: (-s["total"], s["_tb_diff"] if s["_tb_diff"] is not None else 9999))

    # Assign ranks (shared rank for identical total + tiebreaker)
    for i, s in enumerate(scores):
        if i == 0:
            s["rank"] = 1
        else:
            prev = scores[i - 1]
            if s["total"] == prev["total"] and s["_tb_diff"] == prev["_tb_diff"]:
                s["rank"] = prev["rank"]
            else:
                s["rank"] = i + 1

    return scores


# ── Leaderboard sheet ──────────────────────────────────────────────────────────
_GOLD   = PatternFill("solid", fgColor="FFD700")
_SILVER = PatternFill("solid", fgColor="C0C0C0")
_BRONZE = PatternFill("solid", fgColor="CD7F32")
_HEADER_BG = PatternFill("solid", fgColor="1a472a")
_EVEN_BG   = PatternFill("solid", fgColor="F0F4F0")

def _cell(ws, row, col, value, font=None, fill=None, align=None, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font = font
    if fill:   c.fill = fill
    if align:  c.alignment = align
    if num_fmt: c.number_format = num_fmt
    return c


def create_leaderboard_sheet(wb, scores: list, actual: dict, actual_gw: dict,
                              actual_ow, actual_ts, actual_goals):
    """Create (or replace) the Leaderboard sheet."""
    name = "Leaderboard"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(title=name, index=0)

    bold16 = Font(bold=True, size=16, color="1a472a")
    bold11 = Font(bold=True, size=11)
    italic9 = Font(italic=True, size=9)
    centre = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left", vertical="center")
    white_bold = Font(bold=True, color="FFFFFF", size=11)

    # ── Title block ──
    ws.merge_cells("A1:J1")
    _cell(ws, 1, 1, "⚽  World Cup 2026 Predictor  ⚽", font=bold16, align=centre)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ts = datetime.now().strftime("%d %B %Y  %H:%M")
    _cell(ws, 2, 1, f"Last updated: {ts}   |   Matches completed: {len(actual)}", align=centre,
          font=Font(size=10, italic=True))

    gw_text = "  ".join(f"Grp {g}: {w}" for g, w in sorted(actual_gw.items())) or "None yet"
    ws.merge_cells("A3:J3")
    _cell(ws, 3, 1, f"Group leaders: {gw_text}", align=centre, font=Font(size=9))

    if actual_ow or actual_ts:
        ws.merge_cells("A4:J4")
        extra = []
        if actual_ow: extra.append(f"Tournament Winner: {actual_ow}")
        if actual_ts: extra.append(f"Top Scorer: {actual_ts}")
        if actual_goals: extra.append(f"Total Goals: {actual_goals}")
        _cell(ws, 4, 1, "   |   ".join(extra), align=centre, font=Font(size=9))

    # ── Column headers ──
    HDR_ROW = 6
    headers = [
        "Rank", "Name", "Total\nPts", "Match\nPts", "Group\nWins",
        "Winner\nPts", "Scorer\nPts", "Predicted\nWinner", "Predicted\nTop Scorer", "Tiebreaker\n(Goals)"
    ]
    col_widths = [7, 22, 9, 9, 9, 9, 9, 18, 22, 12]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = _cell(ws, HDR_ROW, col, h, font=white_bold,
                  fill=_HEADER_BG, align=Alignment(horizontal="center", vertical="center", wrap_text=True))
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[HDR_ROW].height = 32

    # ── Data rows ──
    for i, s in enumerate(scores):
        row = HDR_ROW + 1 + i

        medal = {1: "🥇", 2: "🥈", 3: "🥉"}.get(s["rank"], "")
        fill = {1: _GOLD, 2: _SILVER, 3: _BRONZE}.get(s["rank"])
        even_fill = _EVEN_BG if (i % 2 == 1 and not fill) else fill

        vals = [
            f"{medal} {s['rank']}",
            s["name"],
            s["total"],
            s["match_pts"],
            s["gw_pts"],
            s["ow_pts"],
            s["ts_pts"],
            s["overall_winner"] or "",
            s["top_scorer"] or "",
            s["tie_breaker"] if s["tie_breaker"] is not None else "",
        ]
        aligns = [centre, left] + [centre] * 8

        for col, (v, a) in enumerate(zip(vals, aligns), 1):
            _cell(ws, row, col, v, fill=even_fill, align=a)

        ws.row_dimensions[row].height = 18

    # ── Scoring key ──
    footer_row = HDR_ROW + len(scores) + 3
    ws.merge_cells(f"A{footer_row}:J{footer_row}")
    _cell(ws, footer_row, 1,
          "Scoring:  Correct Score = 3 pts  |  Correct Result = 1 pt  |  "
          "Group Winner = 3 pts  |  Tournament Winner = 3 pts  |  Top Scorer = 5 pts  |  "
          "Tiebreaker = Total Tournament Goals (closest wins)",
          font=italic9, align=Alignment(horizontal="center", wrap_text=True))
    ws.row_dimensions[footer_row].height = 28

    return ws


# ── Email ──────────────────────────────────────────────────────────────────────
def send_email(config: dict, scores: list, actual: dict, actual_gw: dict,
               actual_ow, actual_ts):
    sender = config.get("email_sender", "").strip()
    password = config.get("email_password", "").strip()
    recipient = config.get("email_recipient", "paulh7741@gmail.com")

    if not sender or not password:
        print("  Email skipped: email_sender / email_password not set in wc_config.json")
        return False

    # Build HTML rows
    rows_html = ""
    for s in scores:
        medal = {1: "🥇 ", 2: "🥈 ", 3: "🥉 "}.get(s["rank"], "")
        bg = {1: "#FFD700", 2: "#C0C0C0", 3: "#CD7F32"}.get(s["rank"], "#ffffff")
        rows_html += f"""
        <tr style="background:{bg}">
          <td style="text-align:center;padding:10px 16px;font-size:16px">{medal}{s['rank']}</td>
          <td style="padding:10px 20px;font-size:16px"><b>{s['name']}</b></td>
          <td style="text-align:center;padding:10px 16px;font-size:16px"><b>{s['total']}</b></td>
        </tr>"""

    html = f"""<html><body style="font-family:Arial,sans-serif;font-size:14px">
<h2 style="color:#1a472a">⚽ World Cup 2026 Predictor – Standings</h2>
<p style="color:#555">Updated: {datetime.now().strftime('%d %B %Y  %H:%M')} &nbsp;|&nbsp; Games completed: {len(actual)}</p>
<table border="1" cellpadding="0" cellspacing="0"
       style="border-collapse:collapse;font-size:15px;min-width:300px">
  <thead>
    <tr style="background:#1a472a;color:white">
      <th style="padding:10px 16px">Rank</th>
      <th style="padding:10px 20px">Name</th>
      <th style="padding:10px 16px">Points</th>
    </tr>
  </thead>
  <tbody>{rows_html}</tbody>
</table>
</body></html>"""

    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"World Cup Predictor – Standings {datetime.now().strftime('%d %b %Y %H:%M')}"
    msg["From"]    = sender
    msg["To"]      = recipient
    msg.attach(MIMEText(html, "html"))

    try:
        with smtplib.SMTP(config.get("smtp_host", "smtp.gmail.com"),
                          config.get("smtp_port", 587)) as smtp:
            smtp.ehlo()
            smtp.starttls()
            smtp.login(sender, password)
            smtp.sendmail(sender, recipient, msg.as_string())
        print(f"  Email sent to {recipient}")
        return True
    except Exception as e:
        print(f"  Email failed: {e}", file=sys.stderr)
        return False


# ── HTML page generator ───────────────────────────────────────────────────────
def generate_html(scores: list, actual: dict, actual_gw: dict) -> str:
    updated = datetime.now().strftime("%d %B %Y  %H:%M")
    rows = ""
    for s in scores:
        medal = {1: "🥇", 2: "🥈", 3: "🥉"}.get(s["rank"], "")
        bg = {1: "#FFD700", 2: "#C0C0C0", 3: "#CD7F32"}.get(s["rank"], "")
        style = f'style="background:{bg}"' if bg else ""
        rows += f"""
        <tr {style}>
          <td>{medal} {s['rank']}</td>
          <td><b>{s['name']}</b></td>
          <td><b>{s['total']}</b></td>
        </tr>"""

    gw_text = " &nbsp;|&nbsp; ".join(
        f"Group {g}: <b>{w}</b>" for g, w in sorted(actual_gw.items())
    ) if actual_gw else "None confirmed yet"

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <meta http-equiv="refresh" content="3600">
  <title>World Cup 2026 Predictor</title>
  <style>
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{ font-family: Arial, sans-serif; background: #f4f4f4; color: #222; padding: 20px; }}
    .container {{ max-width: 520px; margin: 0 auto; }}
    h1 {{ color: #1a472a; text-align: center; font-size: 1.6em; margin-bottom: 6px; }}
    .subtitle {{ text-align: center; color: #555; font-size: 0.9em; margin-bottom: 18px; }}
    .group-leaders {{ background: #fff; border-radius: 8px; padding: 10px 14px;
                      font-size: 0.82em; color: #444; margin-bottom: 16px;
                      border-left: 4px solid #1a472a; }}
    table {{ width: 100%; border-collapse: collapse; background: #fff;
             border-radius: 8px; overflow: hidden;
             box-shadow: 0 2px 8px rgba(0,0,0,0.1); }}
    thead tr {{ background: #1a472a; color: white; }}
    th {{ padding: 12px 16px; text-align: left; font-size: 0.95em; }}
    th:last-child, td:last-child {{ text-align: center; }}
    td {{ padding: 11px 16px; border-bottom: 1px solid #eee; font-size: 0.95em; }}
    tr:last-child td {{ border-bottom: none; }}
    tr:nth-child(even):not([style]) {{ background: #f9f9f9; }}
    .footer {{ text-align: center; font-size: 0.78em; color: #999; margin-top: 14px; }}
    .badge {{ display:inline-block; background:#1a472a; color:white;
              border-radius:12px; padding:2px 10px; font-size:0.8em; }}
  </style>
</head>
<body>
  <div class="container">
    <h1>⚽ World Cup 2026 Predictor</h1>
    <p class="subtitle">Updated: {updated} &nbsp;|&nbsp;
      <span class="badge">{len(actual)} games scored</span></p>
    <div class="group-leaders">
      <b>Group leaders:</b> {gw_text}
    </div>
    <table>
      <thead><tr><th>Rank</th><th>Name</th><th>Points</th></tr></thead>
      <tbody>{rows}</tbody>
    </table>
    <p class="footer">Updates automatically every 2 hours during the tournament.<br>
      Scoring: Correct Score 3pts &bull; Correct Result 1pt &bull;
      Group Winner 3pts &bull; Tournament Winner 3pts &bull; Top Scorer 5pts</p>
  </div>
</body>
</html>"""


# ── GitHub Pages push ──────────────────────────────────────────────────────────
def push_to_github(config: dict, html: str) -> bool:
    token = config.get("github_token", "")
    repo_name = config.get("github_repo", "worldcup-predictor")
    if not token:
        return False
    try:
        from github import Auth, Github, GithubException
        g = Github(auth=Auth.Token(token))
        repo = g.get_user().get_repo(repo_name)

        # Ensure gh-pages branch exists
        try:
            branch = repo.get_branch("gh-pages")
        except GithubException:
            # Create gh-pages from default branch
            sb = repo.get_branch(repo.default_branch)
            repo.create_git_ref(f"refs/heads/gh-pages", sb.commit.sha)

        # Push index.html
        path = "index.html"
        msg = f"Update leaderboard {datetime.now().strftime('%d %b %Y %H:%M')}"
        try:
            existing = repo.get_contents(path, ref="gh-pages")
            repo.update_file(path, msg, html, existing.sha, branch="gh-pages")
        except GithubException:
            repo.create_file(path, msg, html, branch="gh-pages")

        print(f"  Website updated: https://{g.get_user().login}.github.io/{repo_name}/")
        return True
    except Exception as e:
        print(f"  GitHub push failed: {e}", file=sys.stderr)
        return False


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    import argparse

    ap = argparse.ArgumentParser(description="World Cup Predictor – Auto Scorer")
    ap.add_argument("--force-email",  action="store_true", help="Email even if no score changes")
    ap.add_argument("--no-email",     action="store_true", help="Skip email")
    ap.add_argument("--no-api",       action="store_true", help="Use Results sheet only (no API)")
    ap.add_argument("--no-save",      action="store_true", help="Do not save spreadsheet changes")
    args = ap.parse_args()

    # ── Load / create config ──
    if CONFIG_FILE.exists():
        with open(CONFIG_FILE) as f:
            config = {**DEFAULT_CONFIG, **json.load(f)}
    else:
        config = DEFAULT_CONFIG.copy()
        with open(CONFIG_FILE, "w") as f:
            json.dump(config, f, indent=2)
        print(f"Created {CONFIG_FILE}  ← edit this with your API key and email settings.")

    # Allow environment variables to override config (used by GitHub Actions)
    env_map = {
        "WC_FOOTBALL_API_KEY": "football_api_key",
        "WC_EMAIL_SENDER":     "email_sender",
        "WC_EMAIL_PASSWORD":   "email_password",
        "WC_EMAIL_RECIPIENT":  "email_recipient",
        "WC_GITHUB_TOKEN":     "github_token",
        "WC_GITHUB_REPO":      "github_repo",
    }
    for env_key, cfg_key in env_map.items():
        val = os.environ.get(env_key, "").strip()
        if val:
            config[cfg_key] = val

    # ── Load state ──
    state = {}
    if STATE_FILE.exists():
        with open(STATE_FILE) as f:
            state = json.load(f)

    # ── Load workbook ──
    print(f"\nLoading {SPREADSHEET_PATH.name} …")
    wb = openpyxl.load_workbook(str(SPREADSHEET_PATH))

    # ── Parse player sheets ──
    player_sheet_names = [s for s in wb.sheetnames if s not in ("Results", "Leaderboard")]
    players = [parse_player_sheet(wb[s], s) for s in player_sheet_names]

    # Disambiguate duplicate names (case-insensitive) by appending the sheet number
    from collections import Counter
    name_counts = Counter(p["name"].strip().lower() for p in players)
    for p in players:
        if name_counts[p["name"].strip().lower()] > 1:
            sheet_num = p["sheet_name"].split(" - ")[0].strip()
            p["name"] = f"{p['name']} ({sheet_num})"

    print(f"  Parsed {len(players)} player sheets.")

    # ── Parse Results sheet ──
    actual, actual_gw, actual_ow, actual_ts, actual_goals = parse_results_sheet(wb["Results"])
    print(f"  Results sheet: {len(actual)} results, {len(actual_gw)} group winners.")

    # ── Fetch from API ──
    use_api = config.get("use_api", True) and not args.no_api
    if use_api and config.get("football_api_key"):
        print("Fetching live results from football-data.org …")
        api_results, api_gw = fetch_api_results(
            config["football_api_key"], config.get("competition_code", "WC")
        )
        if api_results:
            actual.update(api_results)
        if api_gw:
            actual_gw.update(api_gw)
        if not args.no_save and (api_results or api_gw):
            update_results_sheet(wb["Results"], api_results or {}, api_gw or {})
    elif use_api:
        print("  (No API key in config – using Results sheet only.)")

    # ── Calculate goals tiebreaker from actual results if not in Results sheet ──
    if actual_goals is None and actual:
        actual_goals = sum(s1 + s2 for s1, s2 in actual.values())

    # ── Calculate scores ──
    scores = calculate_scores(players, actual, actual_gw, actual_ow, actual_ts, actual_goals)

    # ── Detect changes ──
    score_snapshot = [(s["name"], s["total"], s["rank"]) for s in scores]
    cur_hash = hashlib.md5(json.dumps(score_snapshot, sort_keys=True).encode()).hexdigest()
    prev_hash = state.get("scores_hash", "")
    changed = cur_hash != prev_hash

    # ── Update spreadsheet ──
    if not args.no_save:
        create_leaderboard_sheet(wb, scores, actual, actual_gw, actual_ow, actual_ts, actual_goals)
        wb.save(str(SPREADSHEET_PATH))
        print(f"  Spreadsheet saved with Leaderboard sheet.")

    # ── Print standings ──
    print(f"\n{'Rank':<5} {'Name':<25} {'Total':>6} {'Match':>6} {'Groups':>7} {'Winner':>7} {'Scorer':>7}")
    print("─" * 65)
    for s in scores:
        print(f"{s['rank']:<5} {s['name']:<25} {s['total']:>6} {s['match_pts']:>6} "
              f"{s['gw_pts']:>7} {s['ow_pts']:>7} {s['ts_pts']:>7}")

    print(f"\nMatches scored: {len(actual)}  |  Group winners known: {len(actual_gw)}")
    if actual_goals:
        print(f"Running total goals: {actual_goals}")

    # ── Website ──
    if config.get("github_token"):
        if changed or args.force_email:
            print("\nPushing to GitHub Pages …")
            html = generate_html(scores, actual, actual_gw)
            push_to_github(config, html)
        else:
            print("\nNo changes – GitHub Pages not updated.")

    # ── Email ──
    send_flag = config.get("send_email_on_change", True) and not args.no_email
    if send_flag and (changed or args.force_email):
        reason = "changes detected" if changed else "forced"
        print(f"Sending email ({reason}) …")
        send_email(config, scores, actual, actual_gw, actual_ow, actual_ts)
    elif send_flag and not changed:
        print("No score changes – email not sent.")

    # ── Save state ──
    state["scores_hash"] = cur_hash
    state["last_run"] = datetime.now().isoformat()
    state["matches_completed"] = len(actual)
    state["group_winners"] = actual_gw
    with open(STATE_FILE, "w") as f:
        json.dump(state, f, indent=2)


if __name__ == "__main__":
    main()
